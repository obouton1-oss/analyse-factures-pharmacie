"""
analyse_consolidee.py — Assemble parser_bo_offrem.py + bdpm.py pour produire
un classeur Excel consolidé : CA par génériqueur, par nature (répertoire /
hybride / remboursement), croisé avec le laboratoire et le mois, et analyse
de la "fuite" vers d'autres génériqueurs pour les molécules disponibles chez
Biogaran.

Usage :
    python3 analyse_consolidee.py fichier1.pdf [fichier2.pdf ...] --data data/ --out rapport.xlsx

Peut aussi être importé et utilisé programmatiquement (voir build_report()).
"""
import argparse
import re
import sys
from pathlib import Path
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from parser_bo_offrem import parse_bo_offrem
from bdpm import BDPM

# ---------------------------------------------------------------------------
# Classification des offres BO-OFFREM -> génériqueur / catégorie
# ---------------------------------------------------------------------------
# Construit à partir de la table "Total Remise en % par fournisseur et par
# offre" du BO-OFFREM (pages de récapitulatif en fin de document), qui donne
# le nom de laboratoire officiel pour chaque libellé d'offre.
# ⚠️ Si de nouvelles offres apparaissent dans de futurs BO-OFFREM et ne sont
# pas dans cette table, elles ressortiront comme "À CLASSIFIER" dans le
# rapport — il suffit alors d'ajouter une ligne ci-dessous.
OFFRE_TO_GENERIQUEUR = {
    # Canal Biogaran Direct (factures unitaires CSP/Movianto) : le parser
    # met toujours offre="BIOGARAN" (pas d'ambiguïté possible sur ce canal),
    # entrée explicite pour ne pas dépendre du filtre heuristique ci-dessous.
    "BIOGARAN": ("BIOGARAN", True),
    # Clé virtuelle, jamais produite par un parser : sert uniquement à faire
    # apparaître "Générique (labo non identifié)" (filet de sécurité BDPM du
    # canal Alliance, cf. _resoudre_marque_alliance) dans les onglots de
    # synthèse (CA par génériqueur, CA type x laboratoire), qui énumèrent
    # les marques à partir de ce dictionnaire.
    "__ALLIANCE_GENERIQUE_NON_IDENTIFIE__": ("Générique (labo non identifié)", True),
    "BIOGARAN Génériques RSF": ("BIOGARAN", True),
    "OFFRE TRINARA RSF": ("EXELTIS", True),
    "EG LABO Génériques RSF": ("EG LABO", True),
    "EG LABO non remise": ("EG LABO", True),
    "SANDOZ Génériques RSF": ("SANDOZ", True),
    "SAWIS Génériques RSF": ("SAWIS", True),
    "MEDAC Génériques RSF": ("MEDAC", True),
    "ARROW Génériques RSF": ("ARROW", True),
    "ARROW non remise": ("ARROW", True),
    "ARROW NR RSF": ("ARROW", True),
    "SUBSTIPHARM non remise": ("SUBSTIPHARM", True),
    "SUN PHARMA Génériques RSF": ("SUN PHARMA", True),
    "SUN PHARMA non remise": ("SUN PHARMA", True),
    "TEVA Génériques RSF": ("TEVA", True),
    "TEVA non remise": ("TEVA", True),
    "VIATRIS Génériques RSF": ("VIATRIS", True),
    "ZENTIVA Génériques RSF": ("ZENTIVA", True),
    "ZENTIVA non remise": ("ZENTIVA", True),
    "SA NR ZENTIVA SORBITOL RSF": ("ZENTIVA", True),
    "OASE NR ACCORD HEALTHCARE RSF": ("ACCORD HEALTHCARE", True),
    "OASE ZYDUS RSF": ("ZYDUS", True),
    "CRISTERS Génériques RSF": ("CRISTERS", True),
    "KRKA Génériques RSF": ("KRKA", True),
    # Offres biosimilaires : gardées comme génériqueur (True) pour ne pas
    # perdre le CA dans le rapport, mais avec une marque distincte pour ne
    # pas les confondre avec les génériques classiques dans les filtres.
    "ACCORD BIOSIMILAIRE RSF": ("ACCORD HEALTHCARE (biosimilaire)", True),
    "FRESENIUS BIOSIMILAIRE RSF": ("FRESENIUS (biosimilaire)", True),
    # Offres non-génériqueurs (dispositifs, nutrition, parapharmacie...) :
    # exclues de l'analyse génériqueur/fuite mais gardées dans le détail.
    "MARQUE CONSEIL PREMIUM RSF": ("DEPOTRADE (dispositifs)", False),
    "NUTRISENS OFFRE_1 GRP RSF": ("NUTRISENS (nutrition)", False),
    "old_NUTRISENS GRP OFFRE 2 RSF": ("NUTRISENS (nutrition)", False),
    "OASE CONVATEC RSF": ("CONVATEC (dispositifs)", False),
    "OASE MOVICOL OTC RSF": ("NORGINE (OTC)", False),
    "DEAL EMBECTA OF3_27,35% RSF": ("EMBECTA (dispositifs)", False),
    "PROMO DAYANG RSF": ("DAYANG (parapharmacie)", False),
    "NESTLE PRIVILEGE AMBASS 2026 OFC RSF": ("NESTLE (nutrition)", False),
    "SIGVARIS CIBLEE 38% RSF": ("SIGVARIS (dispositifs)", False),
    "SIGVARIS DYNAVEN CIBLEE 56% RSF": ("SIGVARIS (dispositifs)", False),
    "COLOPLAST Offre Ciblée 30-15% RSF": ("COLOPLAST (dispositifs)", False),
    "COLOPLAST Offre Ciblée 30% RSF": ("COLOPLAST (dispositifs)", False),
}

MOIS_FR = {
    "JANVIER": 1, "FEVRIER": 2, "FÉVRIER": 2, "MARS": 3, "AVRIL": 4, "MAI": 5,
    "JUIN": 6, "JUILLET": 7, "AOUT": 8, "AOÛT": 8, "SEPTEMBRE": 9,
    "OCTOBRE": 10, "NOVEMBRE": 11, "DECEMBRE": 12, "DÉCEMBRE": 12,
}


def classify_offre(offre_name):
    if offre_name in OFFRE_TO_GENERIQUEUR:
        return OFFRE_TO_GENERIQUEUR[offre_name]
    upper = (offre_name or "").upper()
    for known_offre, (brand, is_gen) in OFFRE_TO_GENERIQUEUR.items():
        prefix = known_offre.split(" Gén")[0].split(" non")[0].split(" NR")[0].split(" RSF")[0]
        if upper.startswith(prefix.upper()):
            return (brand, is_gen)
    return ("À CLASSIFIER", None)


def detect_periode(pdf_path):
    """Extrait le mois/année depuis le nom de fichier ou le contenu du PDF."""
    name = Path(pdf_path).stem.upper()
    for mois_nom, mois_num in MOIS_FR.items():
        m = re.search(rf"{mois_nom}\D*(\d{{4}})", name)
        if m:
            return f"{mois_num:02d}/{m.group(1)}"
    import pdfplumber
    with pdfplumber.open(pdf_path) as pdf:
        text = pdf.pages[0].extract_text() or ""
    for mois_nom, mois_num in MOIS_FR.items():
        m = re.search(rf"{mois_nom}\s+(\d{{4}})", text.upper())
        if m:
            return f"{mois_num:02d}/{m.group(1)}"
    return "Inconnu"


def periode_depuis_date_facture(date_facture):
    """Convertit une date 'JJ/MM/AAAA' (ex: facture Biogaran Direct/Alliance)
    en période 'MM/AAAA', cohérente avec detect_periode() (canal OCP)."""
    if not date_facture or "/" not in date_facture:
        return "Inconnu"
    parts = date_facture.split("/")
    if len(parts) != 3:
        return "Inconnu"
    _, mm, aaaa = parts
    return f"{int(mm):02d}/{aaaa}"


def periode_sort_key(periode):
    """Clé de tri chronologique pour une période 'MM/AAAA' (les périodes
    'Inconnu' sont reléguées à la fin)."""
    if periode == "Inconnu" or "/" not in periode:
        return (9999, 99)
    mm, aaaa = periode.split("/")
    return (int(aaaa), int(mm))


def type_medicament(repertoire):
    """
    Normalise le champ `repertoire` retourné par BDPM.nature() en un type de
    médicament pour la vue croisée type x laboratoire :
      - "Générique (répertoire)" -> inchangé
      - "Princeps" -> inchangé
      - "Hors répertoire (Hybride)" -> "Hybride"
      - "Hors répertoire" -> inchangé
    NB : les biosimilaires ne sont pas distingués (cf. Méthodologie) et
    tombent dans "Hors répertoire" comme le reste des produits hors
    répertoire non hybrides.
    """
    if repertoire == "Hors répertoire (Hybride)":
        return "Hybride"
    return repertoire or "Hors répertoire"


def _enrichir_ligne(r, periode, canal, bdpm, brand, is_gen):
    """Enrichissement commun à tous les canaux : BDPM (répertoire, hybride,
    remboursement, fabricant titulaire). La résolution marque/génériqueur
    (brand, is_gen) est calculée en amont, différemment selon le canal :
    - OCP : classify_offre(offre) — traduit un nom d'offre commerciale
    - Biogaran Direct : toujours ("BIOGARAN", True), sans ambiguïté
    - Alliance : _resoudre_marque_alliance() — mot-clé désignation déjà
      résolu par le parser, filet de sécurité BDPM en second recours
    """
    nat = bdpm.nature(r["cip13"])
    groupe_id, groupe_libelle = bdpm.groupe_generique(r["cip13"])
    hyb = nat.get("hybride") or {}
    return {
        **r,
        "periode": periode,
        "canal": canal,
        "genericqueur": brand,
        "est_genericqueur": is_gen,
        "repertoire": nat["repertoire"],
        "type_medicament": type_medicament(nat["repertoire"]),
        "remboursement": nat["remboursement"],
        "hybride_role": hyb.get("role") or "",
        "groupe_id": groupe_id,
        "groupe_libelle": groupe_libelle,
        "fabricant_titulaire": bdpm.fabricant(r["cip13"]),
    }


def enrich_rows_ocp(pdf_paths, bdpm):
    """Canal OCP : un PDF = un récap mensuel BO-OFFREM (multi-offres)."""
    all_rows = []
    for pdf_path in pdf_paths:
        periode = detect_periode(pdf_path)
        rows, anomalies, _ = parse_bo_offrem(pdf_path)
        if anomalies:
            print(f"⚠️  {len(anomalies)} anomalie(s) sur {pdf_path} :", file=sys.stderr)
            for a in anomalies[:10]:
                print("   ", a, file=sys.stderr)
        for r in rows:
            brand, is_gen = classify_offre(r["offre"])
            all_rows.append(_enrichir_ligne(r, periode, "OCP", bdpm, brand, is_gen))
    return all_rows


def enrich_rows_biogaran_direct(pdf_paths, bdpm):
    """Canal Biogaran Direct : un PDF = une facture de livraison unitaire.
    Génériqueur toujours BIOGARAN (pas d'ambiguïté sur ce canal)."""
    from parser_biogaran_direct import extraire_facture_biogaran_direct
    all_rows = []
    for pdf_path in pdf_paths:
        facture = extraire_facture_biogaran_direct(pdf_path)
        if not facture["lignes"]:
            print(f"⚠️  Aucune ligne produit trouvée dans {pdf_path}", file=sys.stderr)
        periode = periode_depuis_date_facture(facture["date_facture"])
        for r in facture["lignes"]:
            all_rows.append(_enrichir_ligne(r, periode, "Biogaran Direct", bdpm, "BIOGARAN", True))
    return all_rows


# Certaines lignes Alliance n'ont aucun mot-clé de marque dans leur
# désignation (ex. "BISOPROLOL REF CPR 1,25MG 30" pour un Biogaran vendu sans
# suffixe de marque sur cette référence) : parser_alliance.py ne peut alors
# rien détecter dans le texte. Le titulaire officiel BDPM (bdpm.fabricant())
# permet de trancher ces cas avec certitude quand il correspond à un
# génériqueur déjà suivi ailleurs dans le pipeline (OCP, Biogaran Direct) —
# construit à partir des titulaires réels observés sur les lignes "Générique
# (labo non identifié)" du rapport du 09/07/2026 (102 lignes, remontées par
# le pharmacien qui a identifié plusieurs cas Biogaran mal classés qui
# gonflaient à tort l'onglet "Fuite Biogaran"). ⚠️ Le titulaire de l'AMM
# n'est pas toujours strictement le fournisseur commercial (cf. bdpm.py) —
# ne mapper que les libellés confirmés sur des cas réels, pas de règle
# générique automatique (risque de mal classer un cas de co-exploitation).
TITULAIRE_BDPM_TO_GENERIQUEUR = {
    "BIOGARAN": "BIOGARAN",
    "CRISTERS": "CRISTERS",
    "EG LABO - LABORATOIRES EUROGENERICS": "EG LABO",
    "SUN PHARMA FRANCE": "SUN PHARMA",
    "SUN PHARMACEUTICAL INDUSTRIES EUROPE (PAYS BAS)": "SUN PHARMA",
    "VIATRIS": "VIATRIS",
    "EXELTIS SANTE": "EXELTIS",
    "TEVA": "TEVA",
    "TEVA (PAYS-BAS)": "TEVA",
    "ARROW GENERIQUES": "ARROW",
    "SUBSTIPHARM": "SUBSTIPHARM",
}


def _resoudre_marque_alliance(r, bdpm):
    """Résout la marque génériqueur d'une ligne Alliance, par ordre de priorité :
    1. Mot-clé de désignation déjà détecté par parser_alliance.py (reflète la
       marque réellement imprimée sur le produit — la source la plus fiable).
    2. Titulaire officiel BDPM (bdpm.fabricant()), s'il correspond à un
       génériqueur connu (TITULAIRE_BDPM_TO_GENERIQUEUR ci-dessus).
    3. Filet de sécurité final : si le CIP13 est officiellement au répertoire
       générique (source ANSM, indépendante de la marque) sans titulaire
       reconnu, on le signale comme "Générique (labo non identifié)" plutôt
       que de le perdre silencieusement parmi les princeps/dispositifs."""
    if r["offre"]:
        return r["offre"], True
    marque_bdpm = TITULAIRE_BDPM_TO_GENERIQUEUR.get(bdpm.fabricant(r["cip13"]))
    if marque_bdpm:
        return marque_bdpm, True
    nat = bdpm.nature(r["cip13"])
    if nat["repertoire"] == "Générique (répertoire)":
        return "Générique (labo non identifié)", True
    return None, False


def enrich_rows_alliance(pdf_paths, bdpm):
    """Canal Alliance : un PDF = une facture grossiste multi-génériqueurs."""
    from parser_alliance import extraire_facture_alliance
    all_rows = []
    for pdf_path in pdf_paths:
        facture = extraire_facture_alliance(pdf_path)
        if not facture["lignes"]:
            print(f"⚠️  Aucune ligne produit trouvée dans {pdf_path}", file=sys.stderr)
        periode = periode_depuis_date_facture(facture["date_facture"])
        for r in facture["lignes"]:
            brand, is_gen = _resoudre_marque_alliance(r, bdpm)
            all_rows.append(_enrichir_ligne(r, periode, "Alliance", bdpm, brand, is_gen))
    return all_rows


def enrich_rows(ocp_pdfs, bdpm, biogaran_direct_pdfs=None, alliance_pdfs=None):
    """Point d'entrée multi-canal : agrège les lignes enrichies de chaque
    canal fourni. `ocp_pdfs` reste le paramètre historique (positionnel,
    compatible avec les appels existants) ; les autres canaux sont optionnels."""
    all_rows = enrich_rows_ocp(ocp_pdfs, bdpm)
    if biogaran_direct_pdfs:
        all_rows.extend(enrich_rows_biogaran_direct(biogaran_direct_pdfs, bdpm))
    if alliance_pdfs:
        all_rows.extend(enrich_rows_alliance(alliance_pdfs, bdpm))
    return all_rows


def compute_biogaran_catalog(all_rows):
    catalog = set()
    for r in all_rows:
        if r["genericqueur"] == "BIOGARAN" and r["groupe_id"]:
            catalog.add(r["groupe_id"])
    return catalog


def compute_fuite(all_rows, catalog_biogaran):
    fuite = []
    for r in all_rows:
        if (
            r["est_genericqueur"] is True
            and r["genericqueur"] != "BIOGARAN"
            and r["groupe_id"]
            and r["groupe_id"] in catalog_biogaran
        ):
            fuite.append(r)
    return fuite


# ---------------------------------------------------------------------------
# Construction du classeur Excel
# ---------------------------------------------------------------------------
FONT_NAME = "Arial"
NAVY = "1F3864"
RED_LIGHT = "FCE4E4"
GREY_LIGHT = "F2F2F2"
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", start_color=NAVY, end_color=NAVY)
SUBHEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=9)
SUBHEADER_FILL = PatternFill("solid", start_color="2E5395", end_color="2E5395")
TOTAL_FILL = PatternFill("solid", start_color=GREY_LIGHT, end_color=GREY_LIGHT)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16, color=NAVY)
SUBTITLE_FONT = Font(name=FONT_NAME, italic=True, size=10, color="666666")
NORMAL_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, bold=True, size=10)
THIN_BORDER = Border(*(Side(style="thin", color="CCCCCC"),) * 4)
EUR_FMT = '#,##0.00 €;(#,##0.00 €);"-"'
PCT_FMT = '0.0%'

# Palette de couleurs claires, une par génériqueur, réutilisée de façon cohérente
# sur les onglets "CA par génériqueur" et "CA type x laboratoire" (le même
# génériqueur a toujours la même couleur d'un onglet à l'autre).
LABO_PALETTE = [
    "FDEBD3",  # orange clair
    "D6EAF8",  # bleu clair
    "D5F5E3",  # vert clair
    "FCF3CF",  # jaune clair
    "F5EEF8",  # violet clair
    "FADBD8",  # rose clair
    "D1F2EB",  # turquoise clair
    "EBDEF0",  # mauve clair
    "E8F8F5",  # cyan clair
    "FEF9E7",  # crème
    "D4E6F1",  # bleu ciel
    "E9F7EF",  # menthe
    "FDF2E9",  # pêche
    "F4ECF7",  # lavande
]


def build_labo_fill_map(genericqueurs):
    """Associe une PatternFill claire à chaque génériqueur, de façon stable
    (même génériqueur -> même couleur, quel que soit l'onglet)."""
    fills = {}
    for i, gq in enumerate(genericqueurs):
        color = LABO_PALETTE[i % len(LABO_PALETTE)]
        fills[gq] = PatternFill("solid", start_color=color, end_color=color)
    return fills


def style_header_row(ws, row, ncols, col_start=1):
    for c in range(col_start, col_start + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_sheet_detail(wb, all_rows):
    ws = wb.create_sheet("Détail lignes")
    headers = [
        "Période", "CIP13", "Libellé produit", "Laboratoire (génériqueur)",
        "Offre (libellé complet)", "Type de médicament", "Statut répertoire (détail)",
        "Remboursement", "Qté Cdée", "Qté fact", "PPHT Unit (€)", "CA PPHT (€)",
        "Taux Remise", "Remise HT (€)", "Groupe générique (libellé BDPM)",
        "Fabricant (titulaire AMM)", "Rôle hybride (R/H)", "Canal",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for r in all_rows:
        laboratoire = r["genericqueur"] if r["est_genericqueur"] else ""
        ws.append([
            r["periode"], r["cip13"], r["libelle"], laboratoire,
            r["offre"], r["type_medicament"], r["repertoire"], r["remboursement"],
            r["qte_cdee"], r["qte_fact"], r["ppht_unit"], r["ca_ppht"],
            r["taux_remise"] / 100, r["remise_ht"], r["groupe_libelle"] or "",
            r["fabricant_titulaire"] or "", r["hybride_role"], r["canal"],
        ])

    n = len(all_rows) + 1
    for row in range(2, n + 1):
        ws.cell(row=row, column=11).number_format = EUR_FMT
        ws.cell(row=row, column=12).number_format = EUR_FMT
        ws.cell(row=row, column=13).number_format = PCT_FMT
        ws.cell(row=row, column=14).number_format = EUR_FMT
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).font = NORMAL_FONT

    tab = Table(displayName="DetailLignes", ref=f"A1:{get_column_letter(len(headers))}{n}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)

    widths = {"A": 10, "B": 15, "C": 32, "D": 20, "E": 30, "F": 22, "G": 24,
              "H": 16, "I": 9, "J": 9, "K": 12, "L": 12, "M": 10, "N": 12,
              "O": 35, "P": 26, "Q": 12, "R": 16}
    autosize(ws, widths)
    ws.freeze_panes = "A2"
    return ws, n


def build_sheet_genericqueur(wb, n_detail, labo_fills, genericqueurs):
    ws = wb.create_sheet("CA par génériqueur")
    ws["A1"] = "CA par génériqueur — vue consolidée (toutes périodes confondues)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Calculé par formules SUMIF/COUNTIF depuis l'onglet 'Détail lignes'"
    ws["A2"].font = SUBTITLE_FONT

    headers = ["Génériqueur", "Nb lignes", "Qté fact totale", "CA PPHT (€)",
               "Remise HT (€)", "% du CA génériqueurs", "CA Répertoire (€)",
               "CA Hybride (€)", "CA Hors répertoire (€)", "CA Non remboursé (€)"]
    start_row = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=i, value=h)
    style_header_row(ws, start_row, len(headers))

    detail_sheet = "'Détail lignes'"
    # Colonnes de l'onglet Détail lignes (cf. build_sheet_detail) :
    # D=Laboratoire, F=Type de médicament, H=Remboursement, J=Qté fact,
    # L=CA PPHT, N=Remise HT
    col_labo, col_type, col_remb = "D", "F", "H"
    col_qte, col_ca, col_remise = "J", "L", "N"

    first_data_row = start_row + 1
    for i, gq in enumerate(genericqueurs):
        row = first_data_row + i
        rng_labo = f"{detail_sheet}!${col_labo}$2:${col_labo}${n_detail}"
        rng_type = f"{detail_sheet}!${col_type}$2:${col_type}${n_detail}"
        rng_qte = f"{detail_sheet}!${col_qte}$2:${col_qte}${n_detail}"
        rng_ca = f"{detail_sheet}!${col_ca}$2:${col_ca}${n_detail}"
        rng_remise = f"{detail_sheet}!${col_remise}$2:${col_remise}${n_detail}"
        rng_remb = f"{detail_sheet}!${col_remb}$2:${col_remb}${n_detail}"

        ws.cell(row=row, column=1, value=gq)
        ws.cell(row=row, column=2, value=f"=COUNTIF({rng_labo},A{row})")
        ws.cell(row=row, column=3, value=f"=SUMIF({rng_labo},A{row},{rng_qte})")
        ws.cell(row=row, column=4, value=f"=SUMIF({rng_labo},A{row},{rng_ca})")
        ws.cell(row=row, column=5, value=f"=SUMIF({rng_labo},A{row},{rng_remise})")
        ws.cell(row=row, column=7, value=f'=SUMIFS({rng_ca},{rng_labo},A{row},{rng_type},"Générique (répertoire)")')
        ws.cell(row=row, column=8, value=f'=SUMIFS({rng_ca},{rng_labo},A{row},{rng_type},"Hybride")')
        ws.cell(row=row, column=9, value=f'=SUMIFS({rng_ca},{rng_labo},A{row},{rng_type},"Hors répertoire")')
        ws.cell(row=row, column=10, value=f'=SUMIFS({rng_ca},{rng_labo},A{row},{rng_remb},"Non remboursé (ou non renseigné)")')

    last_row = first_data_row + len(genericqueurs) - 1
    total_row = last_row + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = BOLD_FONT
    for col in (2, 3, 4, 5, 7, 8, 9, 10):
        letter = get_column_letter(col)
        c = ws.cell(row=total_row, column=col, value=f"=SUM({letter}{first_data_row}:{letter}{last_row})")
        c.font = BOLD_FONT

    for row in range(first_data_row, total_row + 1):
        for col in (4, 5, 7, 8, 9, 10):
            ws.cell(row=row, column=col).number_format = EUR_FMT
        ws.cell(row=row, column=6, value=f"=D{row}/D${total_row}")
        ws.cell(row=row, column=6).number_format = PCT_FMT
        gq_ligne = ws.cell(row=row, column=1).value
        fill = labo_fills.get(gq_ligne)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=c)
            if cell.font != BOLD_FONT:
                cell.font = NORMAL_FONT
            if fill is not None:
                cell.fill = fill

    widths = {"A": 20, "B": 11, "C": 15, "D": 14, "E": 14, "F": 12, "G": 16, "H": 14, "I": 18, "J": 16}
    autosize(ws, widths)
    return ws


# Canaux détaillés pour le laboratoire Biogaran uniquement (cf. demande
# pharmacien) dans l'onglet "CA type x laboratoire" : une ligne par canal en
# plus de la ligne agrégée "Total", pour voir comment chaque catégorie de
# médicament Biogaran (générique répertoire, hors répertoire, hybride...)
# se répartit entre les 3 flux d'approvisionnement.
CANAUX_BIOGARAN_DETAIL = ["OCP", "Alliance", "Biogaran Direct"]


def build_sheet_type_labo(wb, all_rows, n_detail, periodes, labo_fills):
    """
    Nouvelle vue demandée : CA par type de médicament x par laboratoire,
    avec détail mensuel (une colonne de 4 sous-mesures par période) et un
    bloc "Total annuel" à droite.
    Sous-mesures par période : CA brut (= CA PPHT), Remise HT, CA remisé
    (= CA PPHT − Remise HT, hypothèse à confirmer — voir onglet Méthodologie),
    Taux de remise pondéré (= Remise HT / CA PPHT).

    Pour BIOGARAN uniquement, chaque type de médicament est en plus détaillé
    par canal (OCP / Alliance / Biogaran Direct + une ligne Total) — voir
    CANAUX_BIOGARAN_DETAIL. Les autres laboratoires gardent une seule ligne
    agrégée, comme avant.
    """
    ws = wb.create_sheet("CA type x laboratoire")
    ws["A1"] = "CA par type de médicament et par laboratoire — détail mensuel"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("CA remisé = CA PPHT − Remise HT (hypothèse de calcul à confirmer, "
                "voir onglet Méthodologie). Ne couvre que les lignes génériqueur "
                "(hors dispositifs/nutrition/parapharmacie). Pour Biogaran, détail "
                "par canal (OCP / Alliance / Biogaran Direct) sous chaque type de "
                "médicament ; la ligne « Total » reprend l'agrégat des 3 canaux.")
    ws["A2"].font = SUBTITLE_FONT

    detail_sheet = "'Détail lignes'"
    col_labo, col_type, col_canal = "D", "F", "R"
    col_periode, col_ca, col_remise = "A", "L", "N"
    rng_labo = f"{detail_sheet}!${col_labo}$2:${col_labo}${n_detail}"
    rng_type = f"{detail_sheet}!${col_type}$2:${col_type}${n_detail}"
    rng_canal = f"{detail_sheet}!${col_canal}$2:${col_canal}${n_detail}"
    rng_periode = f"{detail_sheet}!${col_periode}$2:${col_periode}${n_detail}"
    rng_ca = f"{detail_sheet}!${col_ca}$2:${col_ca}${n_detail}"
    rng_remise = f"{detail_sheet}!${col_remise}$2:${col_remise}${n_detail}"

    header_row1, header_row2 = 4, 5
    ws.cell(row=header_row1, column=1, value="Laboratoire")
    ws.cell(row=header_row1, column=2, value="Type de médicament")
    ws.cell(row=header_row1, column=3, value="Canal (Biogaran uniquement)")
    ws.merge_cells(start_row=header_row1, start_column=1, end_row=header_row2, end_column=1)
    ws.merge_cells(start_row=header_row1, start_column=2, end_row=header_row2, end_column=2)
    ws.merge_cells(start_row=header_row1, start_column=3, end_row=header_row2, end_column=3)

    sub_labels = ["CA brut (€)", "Remise (€)", "CA remisé (€)", "Taux remise pondéré"]
    col = 4
    period_col_start = {}
    for periode in periodes:
        period_col_start[periode] = col
        ws.merge_cells(start_row=header_row1, start_column=col, end_row=header_row1, end_column=col + 3)
        ws.cell(row=header_row1, column=col, value=periode)
        for j, lbl in enumerate(sub_labels):
            ws.cell(row=header_row2, column=col + j, value=lbl)
        col += 4
    total_col_start = col
    ws.merge_cells(start_row=header_row1, start_column=col, end_row=header_row1, end_column=col + 3)
    ws.cell(row=header_row1, column=col, value="TOTAL ANNUEL")
    for j, lbl in enumerate(sub_labels):
        ws.cell(row=header_row2, column=col + j, value=lbl)
    n_cols_total = col + 3

    for r in range(header_row1, header_row2 + 1):
        for c in range(1, n_cols_total + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = SUBHEADER_FONT if r == header_row2 else HEADER_FONT
            cell.fill = SUBHEADER_FILL if r == header_row2 else HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

    # Combinaisons (laboratoire, type) réellement présentes dans les données,
    # restreintes aux lignes génériqueur (cohérent avec l'onglet CA par génériqueur).
    combos = sorted({
        (r["genericqueur"], r["type_medicament"])
        for r in all_rows if r["est_genericqueur"] is True
    })

    # Construit la liste finale des lignes à afficher : (labo, type_med, canal).
    # canal = None -> ligne agrégée classique (tous laboratoires hors Biogaran).
    # Pour Biogaran : une ligne par canal de CANAUX_BIOGARAN_DETAIL, puis une
    # ligne canal="Total" qui reprend l'agrégat (identique à l'ancien comportement).
    lignes_a_afficher = []
    for labo, type_med in combos:
        if labo == "BIOGARAN":
            for canal in CANAUX_BIOGARAN_DETAIL:
                lignes_a_afficher.append((labo, type_med, canal))
            lignes_a_afficher.append((labo, type_med, "Total"))
        else:
            lignes_a_afficher.append((labo, type_med, None))

    first_data_row = header_row2 + 1
    for i, (labo, type_med, canal) in enumerate(lignes_a_afficher):
        row = first_data_row + i
        ws.cell(row=row, column=1, value=labo)
        ws.cell(row=row, column=2, value=type_med)
        ws.cell(row=row, column=3, value=canal if canal else "")

        # Critère canal additionnel uniquement sur les lignes de détail par
        # canal (OCP/Alliance/Biogaran Direct) ; les lignes "Total" et les
        # lignes des autres laboratoires restent agrégées tous canaux, comme
        # avant.
        canal_crit = f',{rng_canal},"{canal}"' if canal in CANAUX_BIOGARAN_DETAIL else ""

        for periode in periodes:
            c0 = period_col_start[periode]
            ca_cell = f"{get_column_letter(c0)}{row}"
            remise_cell = f"{get_column_letter(c0 + 1)}{row}"
            ws.cell(row=row, column=c0,
                    value=f'=SUMIFS({rng_ca},{rng_labo},$A{row},{rng_type},$B{row},{rng_periode},"{periode}"{canal_crit})')
            ws.cell(row=row, column=c0 + 1,
                    value=f'=SUMIFS({rng_remise},{rng_labo},$A{row},{rng_type},$B{row},{rng_periode},"{periode}"{canal_crit})')
            ws.cell(row=row, column=c0 + 2, value=f"={ca_cell}-{remise_cell}")
            ws.cell(row=row, column=c0 + 3, value=f"=IFERROR({remise_cell}/{ca_cell},0)")

        c0 = total_col_start
        ca_cell = f"{get_column_letter(c0)}{row}"
        remise_cell = f"{get_column_letter(c0 + 1)}{row}"
        # Total annuel = somme sur TOUTES les lignes de la combinaison, sans filtre
        # de période (donc robuste même si une période n'est pas dans `periodes`,
        # ex. période "Inconnu" faute de PDF exploitable pour un mois donné).
        ws.cell(row=row, column=c0,
                value=f'=SUMIFS({rng_ca},{rng_labo},$A{row},{rng_type},$B{row}{canal_crit})')
        ws.cell(row=row, column=c0 + 1,
                value=f'=SUMIFS({rng_remise},{rng_labo},$A{row},{rng_type},$B{row}{canal_crit})')
        ws.cell(row=row, column=c0 + 2, value=f"={ca_cell}-{remise_cell}")
        ws.cell(row=row, column=c0 + 3, value=f"=IFERROR({remise_cell}/{ca_cell},0)")

    last_row = first_data_row + len(lignes_a_afficher) - 1

    # Ligne TOTAL GÉNÉRAL — exclut explicitement les lignes de détail par canal
    # Biogaran (OCP/Alliance/Biogaran Direct) pour ne pas compter 3 fois le CA
    # déjà inclus dans la ligne "Total" de chaque type de médicament Biogaran.
    total_row = last_row + 1
    ws.cell(row=total_row, column=1, value="TOTAL GÉNÉRAL").font = BOLD_FONT
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    exclusion_canaux = "".join(f',{get_column_letter(3)}{first_data_row}:{get_column_letter(3)}{last_row},"<>{c}"'
                                for c in CANAUX_BIOGARAN_DETAIL)
    for periode in list(periodes) + ["__TOTAL__"]:
        c0 = period_col_start[periode] if periode != "__TOTAL__" else total_col_start
        for offset in (0, 1):  # CA brut, Remise : sommables
            letter = get_column_letter(c0 + offset)
            ws.cell(row=total_row, column=c0 + offset,
                    value=f"=SUMIFS({letter}{first_data_row}:{letter}{last_row}{exclusion_canaux})").font = BOLD_FONT
        ca_cell = f"{get_column_letter(c0)}{total_row}"
        remise_cell = f"{get_column_letter(c0 + 1)}{total_row}"
        ws.cell(row=total_row, column=c0 + 2, value=f"={ca_cell}-{remise_cell}").font = BOLD_FONT
        ws.cell(row=total_row, column=c0 + 3, value=f"=IFERROR({remise_cell}/{ca_cell},0)").font = BOLD_FONT

    # Mise en forme des nombres + police + fond gris sur la ligne total,
    # fond couleur par génériqueur sur les lignes de données. Les lignes de
    # détail par canal Biogaran sont en italique (sous-détail), la ligne
    # "Total" Biogaran repasse en gras (comme l'ancienne ligne unique).
    for row in range(first_data_row, total_row + 1):
        is_total = (row == total_row)
        labo_ligne = ws.cell(row=row, column=1).value
        canal_ligne = ws.cell(row=row, column=3).value
        is_canal_detail = canal_ligne in CANAUX_BIOGARAN_DETAIL
        is_sous_total_biogaran = (canal_ligne == "Total")
        fill = labo_fills.get(labo_ligne) if not is_total else None
        for periode in list(periodes) + ["__TOTAL__"]:
            c0 = period_col_start[periode] if periode != "__TOTAL__" else total_col_start
            ws.cell(row=row, column=c0).number_format = EUR_FMT
            ws.cell(row=row, column=c0 + 1).number_format = EUR_FMT
            ws.cell(row=row, column=c0 + 2).number_format = EUR_FMT
            ws.cell(row=row, column=c0 + 3).number_format = PCT_FMT
        for c in range(1, n_cols_total + 1):
            cell = ws.cell(row=row, column=c)
            if is_total:
                cell.fill = TOTAL_FILL
                if cell.font != BOLD_FONT:
                    cell.font = BOLD_FONT
            else:
                if is_canal_detail:
                    cell.font = Font(name=FONT_NAME, italic=True, size=9, color="555555")
                elif is_sous_total_biogaran:
                    cell.font = BOLD_FONT
                elif cell.font not in (BOLD_FONT,):
                    cell.font = NORMAL_FONT
                if fill is not None:
                    cell.fill = fill

    widths = {"A": 18, "B": 22, "C": 16}
    for c in range(4, n_cols_total + 1):
        widths[get_column_letter(c)] = 13
    autosize(ws, widths)
    ws.freeze_panes = f"D{first_data_row}"
    return ws


def build_sheet_fuite(wb, all_rows, catalog_biogaran):
    ws = wb.create_sheet("Fuite Biogaran")
    ws["A1"] = "Analyse de fuite — génériques disponibles chez Biogaran achetés ailleurs"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("« Disponible chez Biogaran » = constaté par un achat réel sous l'offre "
                "BIOGARAN dans les périodes analysées (voir onglet Méthodologie).")
    ws["A2"].font = SUBTITLE_FONT

    fuite_rows = compute_fuite(all_rows, catalog_biogaran)
    fuite_rows.sort(key=lambda r: -r["ca_ppht"])

    headers = ["Période", "CIP13", "Libellé produit", "Acheté chez", "Groupe générique (BDPM)",
               "Qté fact", "CA PPHT (€)", "Remise HT (€)"]
    start_row = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=i, value=h)
    style_header_row(ws, start_row, len(headers))

    row = start_row + 1
    for r in fuite_rows:
        ws.cell(row=row, column=1, value=r["periode"])
        ws.cell(row=row, column=2, value=r["cip13"])
        ws.cell(row=row, column=3, value=r["libelle"])
        ws.cell(row=row, column=4, value=r["genericqueur"])
        ws.cell(row=row, column=5, value=r["groupe_libelle"] or "")
        ws.cell(row=row, column=6, value=r["qte_fact"])
        ws.cell(row=row, column=7, value=r["ca_ppht"])
        ws.cell(row=row, column=7).number_format = EUR_FMT
        ws.cell(row=row, column=8, value=r["remise_ht"])
        ws.cell(row=row, column=8).number_format = EUR_FMT
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).font = NORMAL_FONT
            ws.cell(row=row, column=c).fill = PatternFill("solid", start_color=RED_LIGHT, end_color=RED_LIGHT)
        row += 1

    last_row = row - 1
    n_fuite = max(0, last_row - start_row)
    if n_fuite > 0:
        tab = Table(displayName="FuiteBiogaran", ref=f"A{start_row}:{get_column_letter(len(headers))}{last_row}")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium11", showRowStripes=True)
        ws.add_table(tab)

        total_row = last_row + 2
        ws.cell(row=total_row, column=3, value="TOTAL FUITE").font = BOLD_FONT
        ws.cell(row=total_row, column=6, value=f"=SUM(F{start_row+1}:F{last_row})").font = BOLD_FONT
        ws.cell(row=total_row, column=7, value=f"=SUM(G{start_row+1}:G{last_row})").font = BOLD_FONT
        ws.cell(row=total_row, column=7).number_format = EUR_FMT
        ws.cell(row=total_row, column=8, value=f"=SUM(H{start_row+1}:H{last_row})").font = BOLD_FONT
        ws.cell(row=total_row, column=8).number_format = EUR_FMT

        # --- Sous-totaux par laboratoire (nouveau), affichés sur la même page,
        # à droite du tableau principal pour rester visibles sans scroller ---
        labos = sorted({r["genericqueur"] for r in fuite_rows})
        sub_col = len(headers) + 2  # 2 colonnes de marge après le tableau principal
        ws.cell(row=start_row, column=sub_col, value="Laboratoire")
        ws.cell(row=start_row, column=sub_col + 1, value="CA brut (€)")
        ws.cell(row=start_row, column=sub_col + 2, value="CA remisé (€)")
        style_header_row(ws, start_row, 3, col_start=sub_col)

        rng_acheteur = f"$D${start_row+1}:$D${last_row}"
        rng_ca_brut = f"$G${start_row+1}:$G${last_row}"
        rng_remise = f"$H${start_row+1}:$H${last_row}"
        sub_row = start_row + 1
        for labo in labos:
            labo_ref = f"{get_column_letter(sub_col)}{sub_row}"
            ws.cell(row=sub_row, column=sub_col, value=labo)
            ws.cell(row=sub_row, column=sub_col + 1,
                    value=f'=SUMIF({rng_acheteur},{labo_ref},{rng_ca_brut})')
            ws.cell(row=sub_row, column=sub_col + 2,
                    value=f'=SUMIF({rng_acheteur},{labo_ref},{rng_ca_brut})-SUMIF({rng_acheteur},{labo_ref},{rng_remise})')
            ws.cell(row=sub_row, column=sub_col + 1).number_format = EUR_FMT
            ws.cell(row=sub_row, column=sub_col + 2).number_format = EUR_FMT
            for c in (sub_col, sub_col + 1, sub_col + 2):
                ws.cell(row=sub_row, column=c).font = NORMAL_FONT
            sub_row += 1
        sub_total_row = sub_row
        ws.cell(row=sub_total_row, column=sub_col, value="TOTAL").font = BOLD_FONT
        for c in (sub_col + 1, sub_col + 2):
            letter = get_column_letter(c)
            ws.cell(row=sub_total_row, column=c,
                    value=f"=SUM({letter}{start_row+1}:{letter}{sub_row-1})").font = BOLD_FONT
            ws.cell(row=sub_total_row, column=c).number_format = EUR_FMT

        autosize(ws, {get_column_letter(sub_col): 20, get_column_letter(sub_col+1): 13, get_column_letter(sub_col+2): 13})
    else:
        ws.cell(row=start_row + 1, column=1, value="Aucune fuite détectée sur la période analysée.")

    widths = {"A": 10, "B": 15, "C": 34, "D": 18, "E": 40, "F": 9, "G": 12, "H": 12}
    autosize(ws, widths)
    ws.freeze_panes = f"A{start_row+1}"
    return ws, n_fuite


def build_sheet_resume(wb, all_rows, n_fuite):
    ws = wb.create_sheet("Résumé", 0)
    ws["A1"] = "Analyse Factures Pharmacie — Récapitulatif"
    ws["A1"].font = TITLE_FONT
    periodes = sorted({r["periode"] for r in all_rows}, key=periode_sort_key)
    ws["A2"] = f"Période(s) analysée(s) : {', '.join(periodes)}"
    ws["A2"].font = SUBTITLE_FONT

    total_ca = sum(r["ca_ppht"] for r in all_rows)
    total_remise = sum(r["remise_ht"] for r in all_rows)
    total_lignes = len(all_rows)
    total_hybrides = sum(1 for r in all_rows if r["type_medicament"] == "Hybride")

    kpis = [
        ("Nombre de lignes produit analysées", total_lignes),
        ("CA PPHT total (€)", total_ca),
        ("Remise HT totale (€)", total_remise),
        ("Nombre de lignes hybrides identifiées", total_hybrides),
        ("Nombre de lignes en fuite Biogaran", n_fuite),
    ]
    row = 4
    for label, val in kpis:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        c = ws.cell(row=row, column=2, value=val)
        if "€" in label:
            c.number_format = EUR_FMT
        c.font = NORMAL_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Voir les onglets :").font = BOLD_FONT
    row += 1
    for txt in [
        "→ 'Détail lignes' : chaque ligne produit enrichie (type, remboursement, hybride, fabricant)",
        "→ 'CA par génériqueur' : synthèse CA/remise par génériqueur, toutes périodes confondues",
        "→ 'CA type x laboratoire' : détail mensuel + total annuel par type de médicament et labo",
        "→ 'Fuite Biogaran' : achats faits ailleurs alors que Biogaran couvre le même générique, "
        "avec sous-totaux par laboratoire",
        "→ 'Méthodologie' : règles et limites de l'analyse (à lire avant interprétation)",
    ]:
        ws.cell(row=row, column=1, value=txt).font = NORMAL_FONT
        row += 1

    autosize(ws, {"A": 60, "B": 18})
    return ws


def build_sheet_methodologie(wb):
    ws = wb.create_sheet("Méthodologie")
    ws["A1"] = "Méthodologie & limites"
    ws["A1"].font = TITLE_FONT

    paragraphs = [
        "",
        "SOURCE DES DONNÉES",
        "- Lignes produit extraites des BO-OFFREM OCP (PDF) via parser_bo_offrem.py, "
        "validé au centime sur plusieurs mois (0 anomalie).",
        "- Enrichissement via les fichiers officiels ANSM (BDPM) : CIS_bdpm.txt, "
        "CIS_CIP_bdpm.txt, CIS_GENER_bdpm.txt, ainsi que le registre ANSM des groupes "
        "hybrides (fic02grp.txt, fic03spe.txt).",
        "",
        "GÉNÉRIQUEUR (fournisseur commercial)",
        "- Déterminé par le nom de l'offre BO-OFFREM (ex : 'BIOGARAN Génériques RSF' "
        "-> BIOGARAN), PAS par le titulaire d'AMM officiel.",
        "- Raison : un même produit peut être commercialisé sous une marque (ex. Biogaran) "
        "alors que le titulaire légal de l'AMM est un autre laboratoire (co-exploitation). "
        "Exemple vérifié sur ce jeu de données : MACROGOL 4000 BIOG a pour titulaire BDPM "
        "'MAYOLY PHARMA FRANCE', pas Biogaran — alors qu'il est bien acheté et facturé "
        "sous marque Biogaran.",
        "",
        "TYPE DE MÉDICAMENT",
        "- 'Générique (répertoire)' / 'Princeps' / 'Hors répertoire' : déterminé via le "
        "fichier CIS_GENER_bdpm.txt (groupes génériques officiels ANSM).",
        "- 'Hybride' : déterminé via le registre ANSM des groupes hybrides. Un hybride "
        "est traité comme une précision de 'Hors répertoire' (les hybrides ne sont pas "
        "au répertoire des génériques) — il ressort comme catégorie 'Hybride' distincte "
        "dans ce rapport plutôt que noyé dans 'Hors répertoire'.",
        "- Remboursement : taux issu de CIS_CIP_bdpm.txt (colonne officielle). "
        "'Non remboursé (ou non renseigné)' regroupe les deux cas faute de pouvoir les "
        "distinguer de façon fiable avec ces seuls fichiers.",
        "",
        "CATÉGORIE NON DISPONIBLE : BIOSIMILAIRES",
        "- Contrairement aux hybrides, les biosimilaires ne sont à ce stade PAS "
        "distingués (aucune source équivalente au registre hybrides n'a encore été "
        "intégrée). Ils tombent actuellement dans 'Hors répertoire'. À investiguer si "
        "besoin — l'ANSM publie une liste de référence des biosimilaires.",
        "",
        "CALCUL DU CA REMISÉ (hypothèse à confirmer)",
        "- CA remisé = CA PPHT − Remise HT. C'est l'hypothèse de calcul retenue pour "
        "l'onglet 'CA type x laboratoire' en l'absence de précision contraire — à "
        "confirmer auprès du pharmacien si un calcul différent (ex. sur la base du "
        "CA PGHT) est attendu.",
        "- Taux de remise pondéré = Remise HT / CA PPHT (donc pondéré par le volume "
        "d'achat, pas une moyenne simple des taux ligne à ligne).",
        "",
        "ANALYSE DE FUITE BIOGARAN",
        "- Principe : un groupe générique (molécule + dosage + forme, code officiel "
        "ANSM) est considéré 'disponible chez Biogaran' s'il a été RÉELLEMENT ACHETÉ "
        "sous l'offre BIOGARAN au moins une fois dans les périodes fournies à ce rapport.",
        "- Ce choix (achat réel plutôt que titulaire BDPM) est déterminant : le titulaire "
        "sous-estime fortement le catalogue Biogaran réel à cause des cas de "
        "co-exploitation (cf. exemple Macrogol ci-dessus).",
        "- Limite : si un mois donné n'a pas été inclus dans l'analyse, un générique "
        "acheté chez Biogaran uniquement ce mois-là n'apparaîtra pas dans le catalogue "
        "de référence. Plus vous fournissez de mois d'historique, plus la détection de "
        "fuite est complète et fiable.",
        "- Sont exclues de l'analyse : les lignes dont l'offre n'est pas un génériqueur "
        "(dispositifs médicaux, nutrition, parapharmacie — ex. Coloplast, Sigvaris, "
        "Nestlé, Convatec) et les produits hors répertoire générique (pas de groupe "
        "générique officiel, donc pas de base de comparaison).",
        "- Les sous-totaux par laboratoire de l'onglet 'Fuite Biogaran' (CA brut / CA "
        "remisé) permettent de voir en un coup d'œil chez qui la fuite se concentre.",
        "",
        "CLASSIFICATION DES OFFRES",
        "- Le mapping offre -> génériqueur est construit à partir de la table "
        "'Total Remise en % par fournisseur et par offre' imprimée en fin de "
        "BO-OFFREM. Si une offre inédite apparaît dans un futur BO-OFFREM sans être "
        "dans ce mapping, elle ressort comme 'À CLASSIFIER' dans le détail — il suffit "
        "d'ajouter une ligne dans OFFRE_TO_GENERIQUEUR (scripts/analyse_consolidee.py).",
    ]
    row = 3
    for p in paragraphs:
        cell = ws.cell(row=row, column=1, value=p)
        if p and p == p.upper() and not p.startswith("-"):
            cell.font = Font(name=FONT_NAME, bold=True, size=11, color=NAVY)
        else:
            cell.font = NORMAL_FONT
        ws.row_dimensions[row].height = 15 if p else 6
        row += 1

    ws.column_dimensions["A"].width = 110
    for r in range(3, row):
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    return ws


def build_report(pdf_paths, data_dir, out_path, biogaran_direct_pdfs=None, alliance_pdfs=None):
    bdpm = BDPM(data_dir)
    all_rows = enrich_rows(pdf_paths, bdpm, biogaran_direct_pdfs=biogaran_direct_pdfs, alliance_pdfs=alliance_pdfs)
    catalog_biogaran = compute_biogaran_catalog(all_rows)
    fuite_rows = compute_fuite(all_rows, catalog_biogaran)
    periodes = sorted({r["periode"] for r in all_rows}, key=periode_sort_key)

    wb = Workbook()
    wb.remove(wb.active)

    genericqueurs = sorted(
        {v[0] for v in OFFRE_TO_GENERIQUEUR.values() if v[1] is True}
        | {r["genericqueur"] for r in all_rows if r["est_genericqueur"] is True and r["genericqueur"]}
    )
    labo_fills = build_labo_fill_map(genericqueurs)

    _, n_detail = build_sheet_detail(wb, all_rows)
    build_sheet_genericqueur(wb, n_detail, labo_fills, genericqueurs)
    build_sheet_type_labo(wb, all_rows, n_detail, periodes, labo_fills)
    _, n_fuite = build_sheet_fuite(wb, all_rows, catalog_biogaran)
    build_sheet_resume(wb, all_rows, n_fuite)
    build_sheet_methodologie(wb)

    wb.save(out_path)
    print(f"Rapport généré : {out_path}")
    print(f"  {len(all_rows)} lignes produit, {len(fuite_rows)} lignes de fuite détectées")
    print(f"  Périodes : {', '.join(periodes)}")
    print(f"  CA PPHT total : {sum(r['ca_ppht'] for r in all_rows):,.2f} €")

    a_classifier = sorted({r["offre"] for r in all_rows if r["genericqueur"] == "À CLASSIFIER"})
    if a_classifier:
        print(f"  ⚠️  {len(a_classifier)} offre(s) non classifiée(s) dans OFFRE_TO_GENERIQUEUR :", file=sys.stderr)
        for o in a_classifier:
            print("   ", o, file=sys.stderr)


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("pdfs", nargs="+", help="Un ou plusieurs BO-OFFREM PDF (canal OCP)")
    ap.add_argument("--data", default="data", help="Dossier des fichiers BDPM")
    ap.add_argument("--out", default="rapport_pharmacie.xlsx", help="Fichier Excel de sortie")
    ap.add_argument("--biogaran-direct", nargs="*", default=None,
                     help="Un ou plusieurs PDF de factures Biogaran Direct (canal Biogaran Direct)")
    ap.add_argument("--alliance", nargs="*", default=None,
                     help="Un ou plusieurs PDF de factures Alliance Healthcare (canal Alliance)")
    args = ap.parse_args()
    build_report(args.pdfs, args.data, args.out,
                  biogaran_direct_pdfs=args.biogaran_direct, alliance_pdfs=args.alliance)
